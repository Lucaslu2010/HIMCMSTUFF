import openpyxl
import numpy as np
import math
import random
import os
from collections import defaultdict


class ExcelReader:
    """Excel文件读取器（支持.xlsx格式）"""

    @staticmethod
    def read_excel_columns(filename, sheet_name=None, columns=['A', 'B', 'C'], header=True):
        """
        从Excel文件读取指定列的数据

        参数:
        filename: Excel文件名
        sheet_name: 工作表名称，如果为None则使用第一个工作表
        columns: 要读取的列字母列表，如['A', 'B', 'C']
        header: 是否跳过第一行（表头）

        返回:
        三维数据点的列表
        """
        if not os.path.exists(filename):
            raise FileNotFoundError(f"文件 {filename} 不存在")

        try:
            # 打开Excel文件
            workbook = openpyxl.load_workbook(filename, data_only=True)

            # 获取工作表
            if sheet_name:
                sheet = workbook[sheet_name]
            else:
                sheet = workbook.active

            data = []
            start_row = 2 if header else 1  # openpyxl行号从1开始

            # 将列字母转换为数字索引
            col_indices = [openpyxl.utils.column_index_from_string(col) for col in columns]

            for row_idx in range(start_row, sheet.max_row + 1):
                row_data = []
                valid_row = True

                for col_idx in col_indices:
                    try:
                        cell_value = sheet.cell(row=row_idx, column=col_idx).value
                        if cell_value is None:
                            valid_row = False
                            break
                        # 尝试转换为浮点数
                        numeric_value = float(cell_value)
                        row_data.append(numeric_value)
                    except (ValueError, TypeError):
                        valid_row = False
                        break

                if valid_row and len(row_data) == len(columns):
                    data.append(row_data)

            print(f"从 {filename} 成功读取 {len(data)} 行数据")
            return data

        except Exception as e:
            raise Exception(f"读取Excel文件时出错: {str(e)}")


class KMeans3D:
    def __init__(self, n_clusters=3, max_iter=300, tol=1e-4, random_state=None):
        """
        初始化K-means聚类器
        """
        self.n_clusters = n_clusters
        self.max_iter = max_iter
        self.tol = tol
        self.random_state = random_state
        self.centroids = None
        self.labels = None
        self.inertia_ = None
        self.history = []

        if random_state is not None:
            random.seed(random_state)
            np.random.seed(random_state)

    def initialize_centroids(self, X):
        """初始化质心 - 随机选择"""
        indices = random.sample(range(len(X)), self.n_clusters)
        return [X[i][:] for i in indices]  # 创建副本

    def euclidean_distance(self, point1, point2):
        """计算欧几里得距离"""
        return math.sqrt(sum((p1 - p2) ** 2 for p1, p2 in zip(point1, point2)))

    def assign_clusters(self, X, centroids):
        """将点分配到最近的簇"""
        labels = []
        for point in X:
            min_dist = float('inf')
            best_cluster = 0
            for i, centroid in enumerate(centroids):
                dist = self.euclidean_distance(point, centroid)
                if dist < min_dist:
                    min_dist = dist
                    best_cluster = i
            labels.append(best_cluster)
        return labels

    def update_centroids(self, X, labels):
        """更新质心位置"""
        new_centroids = []
        cluster_counts = [0] * self.n_clusters
        cluster_sums = [[0.0, 0.0, 0.0] for _ in range(self.n_clusters)]

        # 计算每个簇的和
        for i, point in enumerate(X):
            cluster_id = labels[i]
            cluster_counts[cluster_id] += 1
            for j in range(3):
                cluster_sums[cluster_id][j] += point[j]

        # 计算平均值
        for i in range(self.n_clusters):
            if cluster_counts[i] > 0:
                new_centroid = [s / cluster_counts[i] for s in cluster_sums[i]]
                new_centroids.append(new_centroid)
            else:
                # 如果簇为空，随机选择新点
                new_centroids.append(X[random.randint(0, len(X) - 1)][:])

        return new_centroids

    def calculate_inertia(self, X, labels, centroids):
        """计算簇内平方和"""
        inertia = 0.0
        for i, point in enumerate(X):
            centroid = centroids[labels[i]]
            inertia += self.euclidean_distance(point, centroid) ** 2
        return inertia

    def fit(self, X):
        """训练K-means模型"""
        self.X = X
        print(f"开始K-means聚类，数据点: {len(X)}, 聚类数: {self.n_clusters}")
        print("=" * 50)

        # 初始化质心
        self.centroids = self.initialize_centroids(X)

        for iteration in range(self.max_iter):
            # 分配点到簇
            self.labels = self.assign_clusters(X, self.centroids)

            # 保存旧质心
            old_centroids = self.centroids[:]

            # 更新质心
            self.centroids = self.update_centroids(X, self.labels)

            # 计算最大质心移动
            max_shift = 0.0
            for old, new in zip(old_centroids, self.centroids):
                shift = self.euclidean_distance(old, new)
                max_shift = max(max_shift, shift)

            # 计算惯性
            inertia = self.calculate_inertia(X, self.labels, self.centroids)
            self.history.append({
                'iteration': iteration,
                'inertia': inertia,
                'max_shift': max_shift
            })

            # 打印进度
            if iteration % 10 == 0 or iteration < 5:
                print(f"迭代 {iteration:3d}: 惯性值 = {inertia:10.2f}, 最大移动 = {max_shift:.6f}")

            # 检查收敛
            if max_shift < self.tol:
                print(f"✓ 收敛于第 {iteration} 次迭代")
                break

        self.inertia_ = self.calculate_inertia(X, self.labels, self.centroids)
        print(f"最终惯性值: {self.inertia_:.2f}")
        return self

    def predict(self, new_points):
        """预测新数据点的标签"""
        return self.assign_clusters(new_points, self.centroids)

    def get_cluster_stats(self):
        """获取簇统计信息"""
        stats = {}
        for i in range(self.n_clusters):
            cluster_points = [self.X[j] for j in range(len(self.X)) if self.labels[j] == i]
            if cluster_points:
                points_array = np.array(cluster_points)
                stats[i] = {
                    'size': len(cluster_points),
                    'centroid': self.centroids[i],
                    'mean': np.mean(points_array, axis=0).tolist(),
                    'std': np.std(points_array, axis=0).tolist(),
                    'min': np.min(points_array, axis=0).tolist(),
                    'max': np.max(points_array, axis=0).tolist()
                }
        return stats


class ClusterEvaluator:
    """聚类评估器"""

    @staticmethod
    def silhouette_score(X, labels):
        """计算轮廓系数"""
        n = len(X)
        if len(set(labels)) <= 1:
            return 0.0

        silhouette_scores = []
        for i in range(n):
            point = X[i]
            cluster_id = labels[i]

            # 计算a(i): 到同簇其他点的平均距离
            same_cluster_dists = []
            for j in range(n):
                if labels[j] == cluster_id and j != i:
                    dist = math.sqrt(sum((p1 - p2) ** 2 for p1, p2 in zip(point, X[j])))
                    same_cluster_dists.append(dist)
            a_i = sum(same_cluster_dists) / len(same_cluster_dists) if same_cluster_dists else 0

            # 计算b(i): 到其他簇的最小平均距离
            other_clusters = set(labels) - {cluster_id}
            b_i = float('inf')

            for other_cluster in other_clusters:
                other_dists = []
                for j in range(n):
                    if labels[j] == other_cluster:
                        dist = math.sqrt(sum((p1 - p2) ** 2 for p1, p2 in zip(point, X[j])))
                        other_dists.append(dist)
                if other_dists:
                    avg_dist = sum(other_dists) / len(other_dists)
                    b_i = min(b_i, avg_dist)

            # 计算轮廓系数
            if max(a_i, b_i) > 0:
                s_i = (b_i - a_i) / max(a_i, b_i)
            else:
                s_i = 0
            silhouette_scores.append(s_i)

        return sum(silhouette_scores) / n

    @staticmethod
    def evaluate_clusters(X, labels, centroids):
        """全面评估聚类结果"""
        print("\n" + "=" * 60)
        print("聚类质量评估")
        print("=" * 60)

        # 轮廓系数
        silhouette = ClusterEvaluator.silhouette_score(X, labels)
        print(f"轮廓系数: {silhouette:.4f}")
        if silhouette > 0.7:
            print("  → 优秀聚类 (轮廓系数 > 0.7)")
        elif silhouette > 0.5:
            print("  → 良好聚类 (轮廓系数 > 0.5)")
        elif silhouette > 0.3:
            print("  → 一般聚类")
        else:
            print("  → 聚类效果较差")

        # 惯性值
        inertia = 0
        for i, point in enumerate(X):
            centroid = centroids[labels[i]]
            dist = math.sqrt(sum((p - c) ** 2 for p, c in zip(point, centroid)))
            inertia += dist ** 2
        print(f"簇内平方和: {inertia:.2f}")

        # 簇大小分布
        cluster_sizes = defaultdict(int)
        for label in labels:
            cluster_sizes[label] += 1

        print(f"\n簇分布统计:")
        total_points = len(X)
        for cluster_id in sorted(cluster_sizes.keys()):
            size = cluster_sizes[cluster_id]
            percentage = (size / total_points) * 100
            print(f"  簇 {cluster_id}: {size:4d} 点 ({percentage:5.1f}%)")


class ResultVisualizer:
    """结果可视化器"""

    @staticmethod
    def show_sample_results(X, labels, centroids, sample_size=10):
        """显示样本结果"""
        print("\n" + "=" * 60)
        print("聚类结果样本")
        print("=" * 60)

        print("序号\tX坐标\t\tY坐标\t\tZ坐标\t\t簇标签")
        print("-" * 70)

        for i in range(min(sample_size, len(X))):
            x, y, z = X[i]
            print(f"{i + 1:3d}\t{x:8.3f}\t{y:8.3f}\t{z:8.3f}\t\t{labels[i]}")

    @staticmethod
    def show_centroids(centroids):
        """显示质心信息"""
        print(f"\n最终质心位置:")
        for i, centroid in enumerate(centroids):
            x, y, z = centroid
            print(f"簇 {i}: ({x:8.3f}, {y:8.3f}, {z:8.3f})")

    @staticmethod
    def show_cluster_details(stats):
        """显示簇详细信息"""
        print("\n" + "=" * 60)
        print("簇详细信息")
        print("=" * 60)

        for cluster_id, info in stats.items():
            print(f"\n簇 {cluster_id} (共 {info['size']} 个点):")
            print(f"  质心位置: ({info['centroid'][0]:.3f}, {info['centroid'][1]:.3f}, {info['centroid'][2]:.3f})")
            print(f"  坐标范围: X[{info['min'][0]:.2f}-{info['max'][0]:.2f}], "
                  f"Y[{info['min'][1]:.2f}-{info['max'][1]:.2f}], "
                  f"Z[{info['min'][2]:.2f}-{info['max'][2]:.2f}]")
            print(f"  坐标标准差: ({info['std'][0]:.3f}, {info['std'][1]:.3f}, {info['std'][2]:.3f})")


def elbow_method(X, max_clusters=8):
    """肘部法则确定最佳聚类数"""
    print("正在使用肘部法则确定最佳聚类数...")
    inertias = []
    silhouette_scores = []

    for n in range(1, max_clusters + 1):
        print(f"测试聚类数: {n}", end=" ")
        kmeans = KMeans3D(n_clusters=n, max_iter=100, random_state=42)
        kmeans.fit(X)
        inertias.append(kmeans.inertia_)
        silhouette = ClusterEvaluator.silhouette_score(X, kmeans.labels)
        silhouette_scores.append(silhouette)
        print(f"- 惯性值: {kmeans.inertia_:.2f}, 轮廓系数: {silhouette:.4f}")

    # 寻找最佳聚类数（基于轮廓系数）
    best_n = 1
    best_score = -1
    for n in range(2, max_clusters + 1):  # 从2开始，因为n=1时轮廓系数为0
        if silhouette_scores[n - 1] > best_score:
            best_score = silhouette_scores[n - 1]
            best_n = n

    print(f"\n推荐聚类数: {best_n} (轮廓系数最高: {best_score:.4f})")
    return best_n


def create_sample_excel_file():
    """创建示例Excel文件（.xlsx格式）"""
    try:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "三维数据"

        # 写入表头
        headers = ['X坐标', 'Y坐标', 'Z坐标', '类别']
        for col, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col, value=header)

        # 创建4个簇的示例数据
        np.random.seed(42)
        centers = [
            [1, 1, 1],
            [5, 5, 5],
            [10, 10, 10],
            [15, 15, 15]
        ]

        row_index = 2
        for cluster_id, center in enumerate(centers):
            for _ in range(25):
                point = [coord + np.random.normal(0, 1.5) for coord in center]
                for col, value in enumerate(point, 1):
                    sheet.cell(row=row_index, column=col, value=float(value))
                sheet.cell(row=row_index, column=4, value=f'簇{cluster_id}')
                row_index += 1

        workbook.save('sample_data.xlsx')
        print("已创建示例文件: sample_data.xlsx")
        return 'sample_data.xlsx'

    except Exception as e:
        print(f"创建示例文件时出错: {str(e)}")
        return None


def main():
    """主函数"""
    print("三维K-means聚类程序 - 从Excel读取数据")
    print("=" * 60)

    # 设置Excel文件名
    excel_file = "Kmeans.xlsx"  # 默认文件名

    # 检查文件是否存在，如果不存在则创建示例文件
    if not os.path.exists(excel_file):
        print(f"文件 {excel_file} 不存在")
        print("正在创建示例文件...")
        excel_file = create_sample_excel_file()
        if excel_file is None:
            print("请手动创建Excel文件 (.xlsx格式)")
            print("确保数据在前三列（A、B、C列），第一行可以是表头")
            return

    try:
        # 读取数据（前三列）
        data = ExcelReader.read_excel_columns(excel_file, columns=['A', 'B', 'C'], header=True)

        if len(data) == 0:
            print("错误: 没有读取到有效数据")
            print("请确保Excel文件包含数值数据在前三列")
            return

        print(f"成功读取 {len(data)} 个三维数据点")

        # 显示数据统计
        data_array = np.array(data)
        print(f"数据范围: X[{data_array[:, 0].min():.2f}-{data_array[:, 0].max():.2f}], "
              f"Y[{data_array[:, 1].min():.2f}-{data_array[:, 1].max():.2f}], "
              f"Z[{data_array[:, 2].min():.2f}-{data_array[:, 2].max():.2f}]")

        # 确定最佳聚类数
        optimal_clusters = elbow_method(data, max_clusters=6)

        # 执行最终聚类
        print(f"\n使用最佳聚类数 {optimal_clusters} 进行最终聚类...")
        kmeans = KMeans3D(n_clusters=optimal_clusters, max_iter=300, random_state=42)
        kmeans.fit(data)

        # 评估结果
        ClusterEvaluator.evaluate_clusters(data, kmeans.labels, kmeans.centroids)

        # 显示结果
        ResultVisualizer.show_sample_results(data, kmeans.labels, kmeans.centroids)
        ResultVisualizer.show_centroids(kmeans.centroids)

        # 显示簇统计信息
        stats = kmeans.get_cluster_stats()
        ResultVisualizer.show_cluster_details(stats)

        # 预测示例
        print("\n" + "=" * 60)
        print("新点预测示例")
        print("=" * 60)

        test_points = [
            [data_array[:, 0].mean(), data_array[:, 1].mean(), data_array[:, 2].mean()],
            [data_array[:, 0].min(), data_array[:, 1].min(), data_array[:, 2].min()],
            [data_array[:, 0].max(), data_array[:, 1].max(), data_array[:, 2].max()],
            [0, 0, 0],
            [10, 10, 10]
        ]

        predictions = kmeans.predict(test_points)
        for i, point in enumerate(test_points):
            centroid = kmeans.centroids[predictions[i]]
            distance = math.sqrt(sum((p - c) ** 2 for p, c in zip(point, centroid)))
            print(f"点 {point} → 簇 {predictions[i]} (距离质心: {distance:.3f})")

    except Exception as e:
        print(f"程序执行出错: {str(e)}")
        print("请检查Excel文件格式是否正确")


if __name__ == "__main__":
    # 安装所需库
    print("请确保已安装所需库:")
    print("pip install openpyxl numpy")
    print()

    main()